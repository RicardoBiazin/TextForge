"""Planilha .xlsx: leitura, patch e o que NAO pode ser perdido (etapa 13).

    .\\.venv\\Scripts\\python.exe tests\\teste_xlsx.py

O teste CENTRAL desta suite e' o primeiro: abrir e salvar uma planilha sem
nenhuma edicao devolve o arquivo BYTE A BYTE. E o segundo mais importante e' o do
grafico: editar uma celula e salvar nao pode apagar o que o TextForge nao
entende. Os dois existem porque o caminho obvio -- `openpyxl.load_workbook()`
seguido de `.save()` -- falha nos dois, e falha em SILENCIO: o arquivo continua
valido, so' que sem os graficos.

`openpyxl` e' usado aqui em dois papeis distintos, e vale distinguir: para
CONSTRUIR as fixtures (e' o unico jeito pratico) e para CONFERIR de fora que o
arquivo gravado pelo TextForge continua legivel por outra biblioteca.

A suite PULA inteira se openpyxl nao estiver instalado. Nao precisa de Qt: o
pacote `textforge/planilha/` nao importa PySide6 -- so' o visualizador importa, e
ele nao e' exercitado aqui.
"""

from __future__ import annotations

import datetime as dt
import io
import re
import sys
import zipfile

from ajudantes import (checa, checa_igual, checa_levanta, pasta_temporaria,
                       preparar_qt, pular, resumir, secao)

# ANTES de importar qualquer coisa que arraste Qt: `documento.py` importa
# PySide6, e `preparar_qt` precisa definir QT_QPA_PLATFORM=offscreen primeiro.
TEM_QT = preparar_qt()

try:
    import openpyxl
    from openpyxl.chart import BarChart, Reference
except ImportError:
    openpyxl = None

from textforge.planilha import deteccao, gravador, leitor, valores
from textforge.planilha.folha_xml import (escapar, intervalo_de_sheetdata,
                                          percorrer)
from textforge.planilha.pasta import (TIPO_DATA, TIPO_FORMULA, TIPO_NUMERO,
                                      TIPO_TEXTO)


# ===========================================================================
# Fixtures
# ===========================================================================


def montar_pasta(caminho, *, com_grafico=False, macros=False):
    """Uma planilha pequena com um caso de cada tipo de celula."""
    livro = openpyxl.Workbook()
    aba = livro.active
    aba.title = "Vendas"
    aba["A1"], aba["B1"], aba["C1"] = "Produto", "Valor", "Entrega"
    aba["A2"], aba["B2"] = "Cafe", 12.5
    aba["A3"], aba["B3"] = "Cha", 7
    aba["B4"] = "=SUM(B2:B3)"
    aba["C2"] = dt.datetime(2024, 3, 15)
    aba["C2"].number_format = "DD/MM/YYYY"
    aba["D2"] = True

    segunda = livro.create_sheet("Notas")
    segunda["A1"] = "observacao"
    segunda["A2"] = '  espaco & "aspas" <tag>  '

    if com_grafico:
        grafico = BarChart()
        grafico.add_data(Reference(aba, min_col=2, min_row=1, max_row=3))
        aba.add_chart(grafico, "F5")

    livro.save(caminho)
    if macros:
        # openpyxl nao cria um .xlsm de verdade sem um modelo. Enxertar a parte
        # binaria no pacote basta para o que este teste quer provar: que o
        # gravador COPIA a parte sem olhar para dentro dela.
        _enxertar(caminho, "xl/vbaProject.bin", b"\x00MACRO FALSA\x00")
    return caminho


def _enxertar(caminho, parte: str, conteudo: bytes) -> None:
    """Acrescenta uma parte ao pacote, preservando as demais."""
    original = caminho.read_bytes()
    destino = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(original)) as origem, \
            zipfile.ZipFile(destino, "w", zipfile.ZIP_DEFLATED) as novo:
        for info in origem.infolist():
            novo.writestr(info, origem.read(info.filename))
        novo.writestr(parte, conteudo)
    caminho.write_bytes(destino.getvalue())


def partes_de(dados: bytes) -> dict[str, bytes]:
    with zipfile.ZipFile(io.BytesIO(dados)) as pacote:
        return {i.filename: pacote.read(i.filename) for i in pacote.infolist()}


def primeira(pasta):
    return pasta.folhas[0]


# ===========================================================================
# Requisito 38: sem edicao, o arquivo nao muda
# ===========================================================================


def testar_ida_e_volta() -> None:
    secao("Sem edicao, o arquivo sai IDENTICO")

    with pasta_temporaria() as tmp:
        caminho = montar_pasta(tmp / "intacto.xlsx", com_grafico=True)
        antes = caminho.read_bytes()
        pasta = leitor.abrir(caminho)

        checa(not pasta.alterado, "abrir nao marca nada como alterado")
        checa(pasta.bytes_para_salvar() is pasta.bytes_originais,
              "*** sem edicao, salvar devolve os MESMOS bytes -- nem sequer "
              "recomprimidos (requisito 38) ***")
        checa_igual(pasta.bytes_para_salvar(), antes,
                    "e eles sao byte a byte os do disco")

        # Olhar a planilha nao pode suja-la: trocar de aba, ler celula, pedir a
        # dica de contexto. Se qualquer um desses marcasse `sujas`, a gravacao
        # seguinte reescreveria a aba sem que o usuario tivesse editado nada.
        for folha in pasta.folhas:
            for posicao in list(folha.celulas):
                folha.celula(*posicao)
        checa(not pasta.alterado, "e ler todas as celulas continua nao sujando")


def testar_edicao_minima() -> None:
    secao("Com edicao, so' as partes necessarias mudam")

    with pasta_temporaria() as tmp:
        caminho = montar_pasta(tmp / "editada.xlsx", com_grafico=True)
        antes = partes_de(caminho.read_bytes())

        pasta = leitor.abrir(caminho)
        folha = primeira(pasta)
        checa(pasta.definir(folha, 2, 2, "99,90"), "definir aceita a edicao")
        depois = partes_de(pasta.bytes_para_salvar())

        mudadas = sorted(n for n in antes if antes[n] != depois.get(n))
        checa_igual(mudadas, ["xl/workbook.xml", "xl/worksheets/sheet1.xml"],
                    "*** so' a aba editada e o workbook.xml mudam ***")
        checa_igual(sorted(set(antes) - set(depois)), [],
                    "nenhuma parte desaparece")
        checa_igual(list(antes), list(depois),
                    "*** e a ORDEM das entradas do ZIP e' a mesma (o Excel "
                    "espera [Content_Types].xml primeiro) ***")

        checa(not pasta.definir(folha, 2, 2, "99,90"),
              "redigitar o MESMO valor nao marca nada como alterado")


def testar_grafico_e_macro() -> None:
    secao("O que o TextForge nao entende sobrevive")

    with pasta_temporaria() as tmp:
        caminho = montar_pasta(tmp / "rica.xlsx", com_grafico=True, macros=True)
        antes = partes_de(caminho.read_bytes())

        pasta = leitor.abrir(caminho)
        pasta.definir(primeira(pasta), 2, 1, "Cafe torrado")
        depois = partes_de(pasta.bytes_para_salvar())

        graficos = [n for n in antes if n.startswith("xl/charts/")]
        checa(graficos, f"a fixture tem grafico ({len(graficos)} parte(s))")
        for nome in graficos:
            checa_igual(depois.get(nome), antes[nome],
                        f"*** {nome} sai IDENTICO (openpyxl.save() o "
                        f"descartaria) ***")

        for nome in [n for n in antes if n.startswith("xl/drawings/")]:
            checa_igual(depois.get(nome), antes[nome], f"{nome} intacto")

        checa_igual(depois.get("xl/vbaProject.bin"), antes["xl/vbaProject.bin"],
                    "*** a macro (vbaProject.bin) sai identica ***")
        checa("grafico" in pasta.preservadas and "macro" in pasta.preservadas,
              f"e o usuario e' avisado do que sobrevive: {pasta.preservadas}")


# ===========================================================================
# Fidelidade dos valores
# ===========================================================================


def testar_leitura() -> None:
    secao("Leitura: tipos, formula e data")

    with pasta_temporaria() as tmp:
        pasta = leitor.abrir(montar_pasta(tmp / "ler.xlsx"))
        checa_igual([f.nome for f in pasta.folhas], ["Vendas", "Notas"],
                    "as duas abas aparecem, na ordem do arquivo")
        checa_igual([f.parte for f in pasta.folhas],
                    ["xl/worksheets/sheet1.xml", "xl/worksheets/sheet2.xml"],
                    "*** e cada aba sabe qual ARQUIVO do pacote e' o dela "
                    "(deduzir pelo indice grava na aba errada) ***")
        checa(not pasta.somente_leitura, "a pasta e' gravavel")

        folha = primeira(pasta)
        checa_igual(folha.celula(1, 1).texto, "Produto", "texto (shared string)")
        checa_igual(folha.celula(2, 2).texto, "12.5", "numero com decimal")
        checa_igual(folha.celula(2, 2).tipo, TIPO_NUMERO, "e e' do tipo numero")
        checa_igual(folha.celula(4, 2).texto, "=SUM(B2:B3)",
                    "*** a formula aparece como FORMULA, e nao como o numero "
                    "que o Excel calculou ***")
        checa_igual(folha.celula(4, 2).tipo, TIPO_FORMULA, "e e' do tipo formula")
        checa_igual(folha.celula(2, 3).texto, "15/03/2024",
                    "*** a data aparece como data, e nao como o serial 45366 ***")
        checa_igual(folha.celula(2, 3).tipo, TIPO_DATA, "e e' do tipo data")
        checa_igual(folha.celula(2, 4).texto, "VERDADEIRO", "booleano")
        checa_igual(folha.celula(99, 99).texto, "",
                    "celula fora do preenchido e' vazia, e nao um erro")

        notas = pasta.folhas[1]
        checa_igual(notas.celula(2, 1).texto, '  espaco & "aspas" <tag>  ',
                    "texto com espaco, & e < volta literal")


def testar_gravacao_de_valores() -> None:
    secao("Gravacao: cada tipo volta certo pelo openpyxl")

    with pasta_temporaria() as tmp:
        caminho = montar_pasta(tmp / "tipos.xlsx")
        pasta = leitor.abrir(caminho)
        folha = primeira(pasta)

        pasta.definir(folha, 2, 2, "1.234,56")     # numero pt-BR com milhar
        pasta.definir(folha, 3, 2, "-8")           # negativo
        pasta.definir(folha, 2, 1, '  a & b <c>  ')  # texto que precisa escape
        pasta.definir(folha, 2, 3, "20/12/2025")   # data numa celula de data
        pasta.definir(folha, 4, 2, "=SUM(B2:B3)*2")  # formula nova
        pasta.definir(folha, 7, 1, "linha nova")   # linha que nao existia
        pasta.definir(folha, 3, 5, "coluna nova")  # coluna alem da ultima

        alvo = tmp / "tipos_salva.xlsx"
        alvo.write_bytes(pasta.bytes_para_salvar())

        livro = openpyxl.load_workbook(alvo)
        aba = livro["Vendas"]
        checa_igual(aba["B2"].value, 1234.56,
                    "*** '1.234,56' vira 1234.56, e nao 1.234 (o ponto separa "
                    "MILHAR num numero brasileiro) ***")
        checa_igual(aba["B3"].value, -8, "negativo")
        checa_igual(aba["A2"].value, '  a & b <c>  ',
                    "texto com & < > e espaco nas pontas volta identico")
        checa_igual(aba["C2"].value, dt.datetime(2025, 12, 20),
                    "a data volta como DATA")
        checa_igual(aba["C2"].number_format, "DD/MM/YYYY",
                    "*** e o formato da celula e' preservado: sem isso a data "
                    "apareceria como o numero 46011 ***")
        checa_igual(aba["B4"].value, "=SUM(B2:B3)*2", "a formula nova")
        checa_igual(aba["A7"].value, "linha nova", "linha inexistente foi criada")
        checa_igual(aba["E3"].value, "coluna nova", "coluna alem da ultima")
        checa_igual(aba["A3"].value, "Cha", "e o que nao foi tocado continua la'")
        livro.close()


def testar_esvaziar_celula() -> None:
    secao("Esvaziar uma celula mantem o estilo")

    with pasta_temporaria() as tmp:
        caminho = montar_pasta(tmp / "vazia.xlsx")
        pasta = leitor.abrir(caminho)
        folha = primeira(pasta)
        estilo_antes = _estilo_de(caminho.read_bytes(),
                                  "xl/worksheets/sheet1.xml", "C2")

        pasta.definir(folha, 2, 3, "")
        salvo = pasta.bytes_para_salvar()
        checa_igual(_estilo_de(salvo, "xl/worksheets/sheet1.xml", "C2"),
                    estilo_antes,
                    "*** a celula esvaziada mantem o atributo `s`: apagar o "
                    "estilo junto destruiria a formatacao da coluna ***")

        alvo = tmp / "vazia_salva.xlsx"
        alvo.write_bytes(salvo)
        livro = openpyxl.load_workbook(alvo)
        checa(livro["Vendas"]["C2"].value is None, "e o valor sumiu de fato")
        livro.close()


def _estilo_de(dados: bytes, parte: str, referencia: str) -> str | None:
    """O atributo `s` de uma celula, lido dos bytes crus da aba."""
    aba = partes_de(dados)[parte]
    achado = re.search(rb'<c r="' + referencia.encode() + rb'"([^>]*)>', aba)
    if achado is None:
        return None
    estilo = re.search(rb's="(\d+)"', achado.group(1))
    return estilo.group(1).decode() if estilo else None


# ===========================================================================
# Formulas: recalculo e calcChain
# ===========================================================================


def testar_recalculo() -> None:
    secao("Formula sobrescrita: recalculo e calcChain")

    with pasta_temporaria() as tmp:
        caminho = montar_pasta(tmp / "calc.xlsx")
        _enxertar(caminho, "xl/calcChain.xml",
                  b'<?xml version="1.0"?><calcChain><c r="B4" i="1"/>'
                  b"</calcChain>")

        pasta = leitor.abrir(caminho)
        pasta.definir(primeira(pasta), 4, 2, "42")   # a formula vira literal
        salvo = pasta.bytes_para_salvar()
        depois = partes_de(salvo)

        checa(b'fullCalcOnLoad="1"' in depois["xl/workbook.xml"],
              "*** fullCalcOnLoad=1 entra no workbook.xml: sem ele um total que "
              "dependia da celula continuaria exibindo o numero antigo ***")
        checa("xl/calcChain.xml" not in depois,
              "*** e o calcChain.xml e' REMOVIDO: ele cita uma celula que "
              "deixou de ter formula, e o Excel acusaria 'conteudo ilegivel' ***")
        checa(b"/xl/calcChain.xml" not in depois["[Content_Types].xml"],
              "o Override do calcChain sai do [Content_Types].xml junto")

        alvo = tmp / "calc_salva.xlsx"
        alvo.write_bytes(salvo)
        livro = openpyxl.load_workbook(alvo)
        checa_igual(livro["Vendas"]["B4"].value, 42,
                    "e a celula que era formula agora e' o numero digitado")
        livro.close()


def testar_formula_compartilhada() -> None:
    secao("Formula compartilhada fica travada")

    with pasta_temporaria() as tmp:
        caminho = montar_pasta(tmp / "shared.xlsx")
        _com_formula_compartilhada(caminho)

        pasta = leitor.abrir(caminho)
        folha = primeira(pasta)
        # A fixture deixa `<dimension ref="A1:D4"/>` estreito de proposito: a
        # coluna F fica FORA dele. Em modo streaming o openpyxl acredita no
        # dimension e pararia em D -- as celulas existiriam no arquivo e nao
        # apareceriam na grade. Ver `reset_dimensions` em `leitor._ler_folha`.
        checa_igual(folha.celula(2, 6).tipo, TIPO_FORMULA,
                    "*** celula FORA do <dimension> declarado e' lida assim "
                    "mesmo: o dimension mente com frequencia ***")
        checa(folha.celula(2, 6).travada,
              "*** a celula que GUARDA o texto da formula compartilhada e' "
              "travada: sobrescreve-la apagaria a formula das outras ***")
        checa(not folha.celula(2, 6).editavel,
              "e por isso ela nao aceita edicao na grade")
        checa(folha.celula(2, 2).editavel,
              "enquanto uma celula comum da mesma aba continua editavel")
        checa_igual(folha.celula(2, 6).cache, "25",
                    "o valor que o Excel calculou fica guardado para a dica")


def _com_formula_compartilhada(caminho) -> None:
    """Enxerta uma formula COMPARTILHADA nas linhas 2 e 3 da aba 1.

    A celula F2 guarda o TEXTO da formula (`<f t="shared" si="7" ref="F2:F3">`);
    F3 apenas cita o mesmo `si`, sem texto nenhum. E' o arranjo que o Excel usa
    ao arrastar uma formula por uma coluna inteira, e e' o motivo de F2 nao poder
    ser sobrescrita: o texto dela e' a origem das outras.

    As celulas entram nas linhas que JA' existem, e no fim delas -- inserir uma
    linha fora de ordem produziria um arquivo que nem o Excel nem o leitor
    aceitam, e o teste passaria a medir a fixture em vez do codigo.
    """
    aba = partes_de(caminho.read_bytes())["xl/worksheets/sheet1.xml"]
    inicio, fim, _vazio = intervalo_de_sheetdata(aba)
    novas = {
        2: b'<c r="F2" t="str"><f t="shared" ref="F2:F3" si="7">B2*2</f>'
           b"<v>25</v></c>",
        3: b'<c r="F3" t="str"><f t="shared" si="7"/><v>14</v></c>',
    }

    pedacos, cursor = [aba[:inicio]], inicio
    for linha in percorrer(aba, b"row", inicio, fim):
        numero = int(linha.atributos.get("r", "0"))
        pedacos.append(aba[cursor:linha.inicio_do_fecho])
        if numero in novas:
            pedacos.append(novas[numero])
        cursor = linha.inicio_do_fecho
    pedacos.append(aba[cursor:])
    _trocar_parte(caminho, "xl/worksheets/sheet1.xml", b"".join(pedacos))


def _trocar_parte(caminho, parte: str, conteudo: bytes) -> None:
    original = caminho.read_bytes()
    destino = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(original)) as origem, \
            zipfile.ZipFile(destino, "w", zipfile.ZIP_DEFLATED) as novo:
        for info in origem.infolist():
            novo.writestr(info, conteudo if info.filename == parte
                          else origem.read(info.filename))
    caminho.write_bytes(destino.getvalue())


# ===========================================================================
# Recusar em vez de gravar errado
# ===========================================================================


def testar_recusas() -> None:
    secao("O que abre em somente leitura")

    with pasta_temporaria() as tmp:
        cifrada = tmp / "senha.xlsx"
        cifrada.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64)
        pasta = leitor.abrir(cifrada)
        checa(pasta.somente_leitura, "pasta protegida por senha: somente leitura")
        checa("senha" in pasta.aviso, f"e o motivo diz por que: {pasta.aviso!r}")

        texto = tmp / "nao_e.xlsx"
        texto.write_bytes(b"isto nao e' um ZIP")
        checa(not deteccao.parece_planilha(texto.read_bytes()),
              "*** um .xlsx que nao e' pacote nao entra no modo planilha -- ele "
              "segue pelo caminho normal e vira hexadecimal ***")

        zip_qualquer = tmp / "renomeado.xlsx"
        with zipfile.ZipFile(zip_qualquer, "w") as z:
            z.writestr("leia-me.txt", "sou um zip comum")
        checa(not deteccao.parece_planilha(zip_qualquer.read_bytes()),
              "e um ZIP comum renomeado para .xlsx tambem nao")

        with_dtd = tmp / "dtd.xlsx"
        montar_pasta(with_dtd)
        _trocar_parte(with_dtd, "xl/workbook.xml",
                      b'<?xml version="1.0"?><!DOCTYPE workbook [<!ENTITY a "b">]>'
                      b"<workbook/>")
        pasta = leitor.abrir(with_dtd)
        checa(pasta.somente_leitura and "DTD" in pasta.aviso,
              "*** XML com DTD e' recusado: e' o vetor do billion laughs, e "
              "nenhuma planilha legitima traz um ***")

    secao("E o gravador recusa em vez de tentar")
    with pasta_temporaria() as tmp:
        pasta = leitor.abrir(montar_pasta(tmp / "so_leitura.xlsx"))
        pasta.definir(primeira(pasta), 2, 2, "1")
        pasta.somente_leitura = True
        pasta.aviso = "motivo qualquer"
        checa_levanta(gravador.ErroDeGravacao, pasta.bytes_para_salvar,
                      "gravar uma pasta marcada como somente leitura levanta")


# ===========================================================================
# Salvar duas vezes
# ===========================================================================


def testar_duas_gravacoes() -> None:
    secao("Salvar, editar de novo, salvar")

    with pasta_temporaria() as tmp:
        caminho = montar_pasta(tmp / "duas.xlsx")
        pasta = leitor.abrir(caminho)
        folha = primeira(pasta)

        pasta.definir(folha, 2, 1, "primeira")
        primeiro = pasta.bytes_para_salvar()
        pasta.confirmar_gravacao(primeiro)
        checa(not pasta.alterado, "confirmar_gravacao limpa as celulas sujas")

        pasta.definir(folha, 3, 1, "segunda")
        alvo = tmp / "duas_salva.xlsx"
        alvo.write_bytes(pasta.bytes_para_salvar())

        livro = openpyxl.load_workbook(alvo)
        aba = livro["Vendas"]
        checa_igual(aba["A2"].value, "primeira",
                    "*** a PRIMEIRA edicao continua la' depois da segunda "
                    "gravacao (sem confirmar_gravacao ela seria revertida em "
                    "silencio) ***")
        checa_igual(aba["A3"].value, "segunda", "e a segunda tambem")
        livro.close()


def testar_varias_abas() -> None:
    secao("Edicao em abas diferentes, uma gravacao so'")

    with pasta_temporaria() as tmp:
        caminho = montar_pasta(tmp / "abas.xlsx")
        pasta = leitor.abrir(caminho)
        pasta.definir(pasta.folhas[0], 2, 1, "na primeira")
        pasta.definir(pasta.folhas[1], 3, 1, "na segunda")

        alvo = tmp / "abas_salva.xlsx"
        alvo.write_bytes(pasta.bytes_para_salvar())
        livro = openpyxl.load_workbook(alvo)
        checa_igual(livro["Vendas"]["A2"].value, "na primeira", "aba 1 gravada")
        checa_igual(livro["Notas"]["A3"].value, "na segunda", "aba 2 gravada")
        livro.close()


# ===========================================================================
# Unidades: numeros, datas, referencias, escape
# ===========================================================================


def testar_numeros() -> None:
    secao("Separador decimal deduzido, e nao adivinhado")

    casos = (
        ("1234", 1234.0), ("12,5", 12.5), ("12.5", 12.5),
        ("1.234,56", 1234.56), ("1,234.56", 1234.56),
        ("1.234", 1234.0), ("1,234", 1234.0),
        ("1.2345", 1.2345), ("-8", -8.0), ("+3,5", 3.5),
        ("1e3", 1000.0),
    )
    for texto, esperado in casos:
        checa_igual(valores.ler_numero(texto), esperado,
                    f"ler_numero({texto!r})")

    for texto in ("", "abc", "12a", "-", "1..2,,3", "R$ 10"):
        checa(valores.ler_numero(texto) is None,
              f"ler_numero({texto!r}) nao e' numero")

    checa_igual(valores.numero_como_texto(12.0), "12",
                "inteiro sai sem o '.0' que o Excel mostraria")
    checa_igual(valores.numero_como_texto(12.5), "12.5", "e o decimal com PONTO")


def testar_datas() -> None:
    secao("Numero de serie do Excel")

    checa_igual(valores.serial_de_data(dt.date(2024, 3, 15), False), 45366.0,
                "15/03/2024 e' o serial 45366")
    checa_igual(valores.serial_de_data(dt.date(1900, 3, 1), False), 61.0,
                "*** 01/03/1900 e' 61, e nao 60: o Excel acredita que 1900 foi "
                "bissexto ***")
    checa_igual(valores.serial_de_data(dt.datetime(2024, 3, 15, 12, 0), False),
                45366.5, "meio-dia e' meia unidade")
    checa_levanta(ValueError, valores.serial_de_data, "data antes de 01/03/1900 "
                  "e' recusada em vez de gravar um dia errado",
                  dt.date(1899, 12, 31), False)

    checa_igual(valores.ler_data("15/03/2024"), dt.datetime(2024, 3, 15),
                "le' o formato pt-BR")
    checa_igual(valores.ler_data("2024-03-15"), dt.datetime(2024, 3, 15),
                "e o ISO")
    checa_igual(valores.ler_data("03/04/2024"), dt.datetime(2024, 4, 3),
                "*** 03/04 e' 3 de ABRIL: o formato pt-BR e' tentado antes do "
                "americano ***")
    checa(valores.ler_data("nao e' data") is None, "e o que nao e' data da' None")


def testar_referencias() -> None:
    secao("Referencias de celula")

    for coluna, letra in ((1, "A"), (26, "Z"), (27, "AA"), (52, "AZ"),
                          (703, "AAA")):
        checa_igual(valores.letra_de_coluna(coluna), letra,
                    f"coluna {coluna} e' {letra}")
        checa_igual(valores.coluna_de_letra(letra), coluna,
                    f"e {letra} volta a ser {coluna}")

    checa_igual(valores.de_referencia("B7"), (7, 2), "B7 e' linha 7, coluna 2")
    checa_igual(valores.de_referencia("$AA$100"), (100, 27),
                "referencia absoluta tambem")
    checa(valores.de_referencia("B7:C9") is None, "faixa nao e' celula")


def testar_escape() -> None:
    secao("Escape de XML")

    checa_igual(escapar('a & b < c > d'), "a &amp; b &lt; c &gt; d",
                "*** o '&' e' escapado PRIMEIRO: na ordem inversa, um '<' "
                "recem-escrito viraria '&amp;lt;' ***")
    checa_igual(escapar("sem nada"), "sem nada", "texto comum passa igual")


# ===========================================================================
# O caminho de verdade: Documento.abrir -> editar -> salvar
# ===========================================================================


def testar_documento() -> None:
    secao("Documento em MODO PLANILHA")

    from textforge.documento import MODO_PLANILHA, Documento
    from textforge.linguagens import carregar_embutidos

    # O `app.py` faz isto na partida; num teste o registro comeca vazio, e sem
    # os provedores toda linguagem seria resolvida como "Texto".
    carregar_embutidos()

    with pasta_temporaria() as tmp:
        caminho = montar_pasta(tmp / "doc.xlsx", com_grafico=True)
        antes = caminho.read_bytes()

        doc = Documento.abrir(caminho)
        checa_igual(doc.modo, MODO_PLANILHA,
                    "*** o .xlsx entra em MODO PLANILHA, e nao em MODO HEX -- "
                    "sem o desvio a assinatura PK do ZIP o mandaria para o "
                    "visualizador hexadecimal ***")
        checa(doc.planilha is not None, "e o documento tem a Pasta")
        checa(not doc.binario,
              "`binario` fica False: o conteudo E' exibido, so' que em grade")
        checa(not doc.somente_leitura, "e a planilha e' gravavel")
        checa_igual(doc.nome_da_linguagem, "Planilha",
                    "a barra de status diz Planilha")
        checa_igual(doc.texto(), "",
                    "*** o QTextDocument fica VAZIO: quem tem o conteudo e' a "
                    "Pasta, e nao o editor de texto ***")

        checa_igual(doc.bytes_para_salvar(), antes,
                    "*** salvar sem editar devolve o arquivo identico ***")

        doc.planilha.definir(doc.planilha.folhas[0], 2, 2, "77")
        doc.salvar()
        checa(not doc.modificado, "depois de salvar, o documento nao esta' sujo")
        checa(not doc.planilha.alterado,
              "e a Pasta foi promovida (salvar de novo nao reverte a edicao)")

        livro = openpyxl.load_workbook(caminho)
        checa_igual(livro["Vendas"]["B2"].value, 77, "o disco tem o valor novo")
        checa(len(livro["Vendas"]._charts) == 1,
              "*** e o grafico continua no arquivo depois de salvar pelo "
              "Documento ***")
        livro.close()

        doc.recarregar()
        checa_igual(doc.planilha.folhas[0].celula(2, 2).texto, "77",
                    "recarregar traz a Pasta do disco")
        checa_levanta(ValueError, doc.reabrir_como,
                      "reabrir com outra codificacao e' recusado: um .xlsx nao "
                      "tem codificacao para trocar", "cp1252")


def testar_extensao_mentirosa() -> None:
    secao("Arquivo que so' PARECE planilha pela extensao")

    from textforge.documento import MODO_PLANILHA, Documento

    with pasta_temporaria() as tmp:
        falso = tmp / "renomeado.xlsx"
        with zipfile.ZipFile(falso, "w") as pacote:
            pacote.writestr("leia-me.txt", "sou um zip comum")

        doc = Documento.abrir(falso)
        checa(doc.modo != MODO_PLANILHA,
              "*** um ZIP comum renomeado para .xlsx NAO vira planilha: a "
              "deteccao olha o conteudo, e nao so' a extensao ***")
        checa(doc.binario, "ele segue pelo caminho normal e e' binario")


# ===========================================================================


def main() -> int:
    if openpyxl is None:
        return pular("openpyxl nao esta' instalado "
                     "(.venv\\Scripts\\python.exe -m pip install openpyxl)")

    testar_ida_e_volta()
    testar_edicao_minima()
    testar_grafico_e_macro()
    testar_leitura()
    testar_gravacao_de_valores()
    testar_esvaziar_celula()
    testar_recalculo()
    testar_formula_compartilhada()
    testar_recusas()
    testar_duas_gravacoes()
    testar_varias_abas()
    testar_numeros()
    testar_datas()
    testar_referencias()
    testar_escape()
    if TEM_QT:
        testar_documento()
        testar_extensao_mentirosa()
    else:
        print("\n[PySide6 ausente: a parte de Documento foi pulada]")
    return resumir()


if __name__ == "__main__":
    sys.exit(main())
