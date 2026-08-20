"""Leitura dos VALORES da planilha. E' o unico lugar que importa openpyxl.

A divisao com o `gravador.py` e' deliberada e nao deve ser afrouxada: openpyxl e'
excelente para LER -- resolve sharedStrings, strings inline, texto rico, formatos
de data e formulas, e faz isso em streaming --, e e' destrutivo para ESCREVER,
porque `save()` reconstroi o pacote inteiro a partir do que ele entendeu e
descarta o que nao entendeu.

Duas escolhas de exibicao:

**Formula aparece como formula.** Com `data_only=False` a celula de `=SOMA(A1:A9)`
mostra a formula, e nao o numero. E' o que um editor de arquivos deve fazer: ele
mostra o arquivo, nao a interpretacao dele. O valor que o Excel calculou fica na
dica de contexto -- lido dos bytes crus, e nao de uma segunda passagem com
`data_only=True`, que custaria abrir a pasta duas vezes.

**Data aparece como data.** O `45366` guardado na celula so' e' 15/03/2024 por
causa do formato numerico, que mora no estilo. Mostrar o numero de serie seria
tecnicamente fiel e praticamente inutil.
"""

from __future__ import annotations

import datetime as dt
import io
import pathlib
import warnings
import zipfile

from textforge import arquivos, log_interno
from textforge.planilha import deteccao, folha_xml, valores
from textforge.planilha.pasta import (Celula, Folha, Pasta, TIPO_BOOL,
                                      TIPO_DATA, TIPO_ERRO, TIPO_FORMULA,
                                      TIPO_NUMERO, TIPO_TEXTO)

log = log_interno.obter(__name__)

#: Teto de celulas PREENCHIDAS carregadas. Passando disso a pasta abre em
#: somente leitura com o que coube: exibir uma planilha pela metade e' util,
#: grava-la pela metade apagaria o resto.
TETO_DE_CELULAS = 500_000

#: Aba maior que isto nao tem as formulas compartilhadas rastreadas -- a
#: varredura custaria mais que o beneficio. A aba inteira vira somente leitura,
#: que e' o lado seguro.
TETO_DA_VARREDURA = 64 * 1024 * 1024


def abrir(caminho, cfg: dict | None = None,
          dados: bytes | None = None) -> Pasta:
    """Le' a planilha inteira. Nunca levanta por conteudo: o motivo vem no aviso.

    `dados` evita a segunda leitura do disco: quem chama ja' precisou dos bytes
    para saber que o arquivo era uma planilha.
    """
    cfg = cfg or {}
    alvo = pathlib.Path(caminho)
    if dados is None:
        dados = arquivos.ler_bytes(alvo)
    pasta = Pasta(alvo, dados)

    pacote = deteccao.inspecionar(dados)
    pasta.data1904 = pacote.data1904
    pasta.preservadas = list(pacote.preservadas)
    if pacote.motivo_somente_leitura:
        pasta.somente_leitura = True
        pasta.aviso = pacote.motivo_somente_leitura

    for aba in pacote.abas:
        folha = Folha(nome=aba.nome, parte=aba.parte, oculta=aba.oculta)
        if not aba.parte:
            folha.motivo_somente_leitura = (
                "aba sem arquivo declarado em xl/_rels/workbook.xml.rels")
        pasta.folhas.append(folha)

    if pacote.e_planilha:
        try:
            _carregar_valores(pasta, dados, int(
                cfg.get("limite_celulas_planilha", TETO_DE_CELULAS)))
        except Exception as exc:                # noqa: BLE001 - arquivo externo
            log.warning("falha ao ler os valores de %s: %s", alvo, exc)
            pasta.somente_leitura = True
            pasta.aviso = pasta.aviso or f"planilha ilegivel: {exc}"

    log.info("planilha aberta %s (%d bytes, %d aba(s), %s)", alvo, len(dados),
             len(pasta.folhas), pasta.aviso or "gravavel")
    return pasta


# ---------------------------------------------------------------------------
# Valores
# ---------------------------------------------------------------------------


def _carregar_valores(pasta: Pasta, dados: bytes, teto: int) -> None:
    import openpyxl

    with warnings.catch_warnings():
        # openpyxl avisa sobre extensoes que nao sabe ler (validacao de dados,
        # formatacao condicional nova). Aqui isso e' esperado: essas partes NAO
        # sao lidas nem regravadas -- sao copiadas inteiras pelo gravador.
        warnings.simplefilter("ignore")
        pasta_xl = openpyxl.load_workbook(io.BytesIO(dados), read_only=True,
                                          data_only=False, keep_links=False)
        try:
            restante = teto
            for folha in pasta.folhas:
                if folha.nome not in pasta_xl.sheetnames:
                    folha.motivo_somente_leitura = "aba nao encontrada no pacote"
                    continue
                restante = _ler_folha(folha, pasta_xl[folha.nome], restante)
                if restante <= 0:
                    pasta.somente_leitura = True
                    pasta.aviso = (f"planilha com mais de {teto:,} celulas: "
                                   "somente leitura".replace(",", "."))
                    break
        finally:
            pasta_xl.close()

    _marcar_formulas(pasta, dados)


def _ler_folha(folha: Folha, aba, restante: int) -> int:
    """Preenche `folha.celulas`. Devolve quantas celulas ainda cabem no teto.

    `reset_dimensions()` NAO e' opcional. Em modo streaming o openpyxl acredita
    no `<dimension ref="A1:D4"/>` declarado na aba e para de ler nos limites
    dele. E `dimension` mente com frequencia: programas que geram planilha o
    escrevem estreito demais, e uma coluna inteira de dados que existe no
    arquivo simplesmente nao apareceria na grade. Sem isto, o editor mostraria
    menos do que o arquivo tem -- e nada avisaria.
    """
    reiniciar = getattr(aba, "reset_dimensions", None)
    if callable(reiniciar):
        reiniciar()
    for linha in aba.iter_rows():
        for celula in linha:
            if celula.value is None:
                continue
            if restante <= 0:
                folha.motivo_somente_leitura = "aba lida so' em parte"
                return 0
            restante -= 1
            posicao = (celula.row, celula.column)
            folha.celulas[posicao] = _converter(celula)
            folha.linhas = max(folha.linhas, celula.row)
            folha.colunas = max(folha.colunas, celula.column)
    return restante


def _converter(celula) -> Celula:
    """Uma celula do openpyxl no que a grade precisa."""
    valor = celula.value
    tipo_ooxml = getattr(celula, "data_type", "")

    if tipo_ooxml == "f" or (isinstance(valor, str) and valor.startswith("=")):
        texto = str(valor)
        return Celula(texto=texto, tipo=TIPO_FORMULA, ordenacao=texto.lower())

    if tipo_ooxml == "e":
        texto = str(valor)
        # Um erro do Excel (#REF!, #DIV/0!) fica visivel e travado: nao ha' texto
        # que o usuario possa digitar que signifique "um erro", e aceitar a
        # edicao o transformaria numa string parecida com um erro.
        return Celula(texto=texto, tipo=TIPO_ERRO, ordenacao=texto,
                      travada=True)

    if isinstance(valor, bool):
        texto = "VERDADEIRO" if valor else "FALSO"
        return Celula(texto=texto, tipo=TIPO_BOOL, ordenacao=texto.lower())

    if isinstance(valor, (dt.datetime, dt.date, dt.time, dt.timedelta)):
        texto = valores.data_como_texto(valor)
        try:
            chave: float | str = valores.serial_de_data(valor, False)
        except (ValueError, TypeError):
            chave = texto
        return Celula(texto=texto, tipo=TIPO_DATA, ordenacao=chave)

    if isinstance(valor, (int, float)):
        return Celula(texto=valores.numero_como_texto(valor), tipo=TIPO_NUMERO,
                      ordenacao=float(valor))

    texto = str(valor)
    return Celula(texto=texto, tipo=TIPO_TEXTO, ordenacao=texto.lower())


# ---------------------------------------------------------------------------
# Formulas: o valor em cache e a formula COMPARTILHADA
# ---------------------------------------------------------------------------


def _marcar_formulas(pasta: Pasta, dados: bytes) -> None:
    """Guarda o valor calculado e trava as celulas de formula compartilhada.

    Uma formula compartilhada (`<f t="shared" si="3" ref="B2:B99">`) tem o texto
    numa celula so'; as outras 97 apenas citam o `si`. Sobrescrever a celula que
    guarda o texto apagaria a formula de todas as demais -- por isso ela e'
    travada, em vez de o pacote inteiro virar somente leitura.
    """
    formulas = [f for f in pasta.folhas
                if f.parte and any(c.tipo == TIPO_FORMULA
                                   for c in f.celulas.values())]
    if not formulas:
        return

    try:
        with zipfile.ZipFile(io.BytesIO(dados)) as zip_:
            tamanhos = {info.filename: info.file_size for info in zip_.infolist()}
            for folha in formulas:
                if tamanhos.get(folha.parte, 0) > TETO_DA_VARREDURA:
                    folha.motivo_somente_leitura = (
                        "aba grande demais para conferir formulas "
                        "compartilhadas")
                    continue
                try:
                    _varrer_aba(folha, zip_.read(folha.parte))
                except (KeyError, ValueError) as exc:
                    log.warning("aba %r nao varrida: %s", folha.nome, exc)
                    folha.motivo_somente_leitura = f"aba ilegivel: {exc}"
    except (zipfile.BadZipFile, OSError) as exc:
        log.warning("varredura de formulas abortada: %s", exc)


def _varrer_aba(folha: Folha, dados: bytes) -> None:
    inicio, fim, vazio = folha_xml.intervalo_de_sheetdata(dados)
    if vazio:
        return
    for linha in folha_xml.percorrer(dados, b"row", inicio, fim):
        for celula in folha_xml.percorrer(dados, b"c", linha.fim_da_abertura,
                                          linha.inicio_do_fecho):
            referencia = celula.atributos.get("r", "")
            posicao = valores.de_referencia(referencia) if referencia else None
            if posicao is None:
                continue
            alvo = folha.celulas.get(posicao)
            if alvo is None or alvo.tipo != TIPO_FORMULA:
                continue
            for formula in folha_xml.percorrer(dados, b"f",
                                               celula.fim_da_abertura,
                                               celula.inicio_do_fecho):
                if formula.atributos.get("t") == "shared":
                    alvo.travada = True
                break
            for calculado in folha_xml.percorrer(dados, b"v",
                                                 celula.fim_da_abertura,
                                                 celula.inicio_do_fecho):
                alvo.cache = folha_xml.desescapar(
                    dados[calculado.fim_da_abertura:calculado.inicio_do_fecho]
                    .decode("utf-8", "replace"))
                break
